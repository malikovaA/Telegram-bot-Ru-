from .sqlalch import session,Test_name, Test_answer, Test_question
from openpyxl import load_workbook, Workbook
# import datetime

# import telebot
# bot = telebot.TeleBot(token)
#
def export_test_class(results):
    wb = Workbook()
    worksheet = wb['Sheet']
    column_1 = [str(i[0])+' баллов' for i in results]
    column_2 = [i[1] for i in results]
    column_3 = [str(i[2]) for i in results]
    column_4 = [i[3] for i in results]
    rows_for_export = [[a,b,c,d] for a,b,c,d in zip(column_1, column_2, column_3, column_4)]
    for row in rows_for_export:
        worksheet.append(row)
    file_name = f'db_pack/files/export_class_.xlsx'
    wb.save(file_name)
    return file_name

def export_test_theme(results):
    wb = Workbook()
    worksheet = wb['Sheet']
    column_1 = [str(i[0])+' баллов' for i in results]
    column_2 = [i[1] for i in results]
    column_3 = [i[2]+5 for i in results]
    column_4 = [i[4] for i in results]
    rows_for_export = [[a,b,c,d] for a,b,c,d in zip(column_1, column_2, column_3, column_4)]
    for row in rows_for_export:
        worksheet.append(row)
    file_name = f'db_pack/files/export_theme_.xlsx'
    wb.save(file_name)
    return file_name




def get_or_create(session, model, **kwargs):
    instance = session.query(model).filter_by(**kwargs).first()
    if instance:

        return instance
    else:
        instance = model(**kwargs)
        session.add(instance)
        session.commit()
        return instance


'''В функцию передаётся название файла (получаем из метода telebot по полученному файлу)'''
def import_test(doc_name):
    wb = load_workbook(filename=doc_name)
    sheet = wb.active
    new_test = Test_name(title=doc_name)
    '''Создается новый экземпляр теста со значением имени файла (название теста)'''
    session.add(new_test)
    session.commit()
    '''Чтение каждой строки'''
    for row in sheet:
        '''Внутри каждой строки происходит чтение каждой ячейки'''
        new_question = Test_question()
        for count, cell in enumerate(row):
            '''В первой ячейке хранится вопрос'''
            if count == 0:
                new_question = Test_question(content=cell.value, test_id=new_test.id)
                session.add(new_question)
                session.commit()
                '''Следующие 3 ячейки содержат неправильные ответы, значение right по умолчанию false'''
            elif 0 < count < 4 :
                new_answer = Test_answer(content=cell.value, test_q_id=new_question.id)
                session.add(new_answer)
                session.commit()
                '''В пятой ячейке всегда должен быть правильный ответ, его записываем с параметром right 1'''
            elif count == 4:
                new_answer = Test_answer(content=cell.value, right=1, test_q_id=new_question.id)
                session.add(new_answer)
                session.commit()
    return 'Тест успешно импортирован'

# import_test('test.xlsx')
